import streamlit as st




def god(STATE):
    import requests
    from bs4 import BeautifulSoup
    import pandas as pd
    import openpyxl
    import os

    output_directory = f"D:/{STATE}"

    if not os.path.exists(output_directory):
        os.makedirs(output_directory)
    
    def arbaj(STATE):
        STATE_URL = f"https://www.censusindia2011.com/{STATE}-population.html"

        def fetch_html(url):
            try:
                response = requests.get(url)
                if response.status_code == 200:
                    return response.text
                else:
                    print(f"Failed to fetch HTML. Status code: {response.status_code}")
                    return None
            except Exception as e:
                print(f"An error occurred: {str(e)}")
                return None

        url = STATE_URL
        html = fetch_html(url)

        def extract_district_names(html):
            district_names = []
            soup = BeautifulSoup(html, 'html.parser')
            tables = soup.find_all('table')
            for table in tables:
                headers = [header.text.strip() for header in table.find_all('th')]
                if 'District' in headers:
                    rows = table.find_all('tr')
                    for row in rows[1:]:
                        cells = row.find_all('td')
                        district_name = cells[0].text.strip()
                        district_names.append(district_name)
            return district_names

        district_names = extract_district_names(html)
        return district_names

    districts = arbaj(STATE)

    def format_district_name(district_name):
        if ' ' in district_name:
            return district_name.replace(" ", "-")
        else:
            return district_name

    for district in districts:
        formatted_district_name = format_district_name(district)
        DISTRICT_URL = f"https://www.censusindia2011.com/{STATE}/{formatted_district_name}-population.html"
        def fetch_html(url):
            try:
                response = requests.get(url)
                if response.status_code == 200:
                    return response.text
                else:
                    print(f"Failed to fetch HTML. Status code: {response.status_code}")
                    return None
            except Exception as e:
                print(f"An error occurred: {str(e)}")
                return None
        html = fetch_html(DISTRICT_URL)
        wb = openpyxl.Workbook()

        def extract_taluka_names(html):
            taluka_names = []
            soup = BeautifulSoup(html, 'html.parser')
            tables = soup.find_all('table')

            for table in tables:
                headers = [header.text.strip() for header in table.find_all('th')]
                taluka_headers = ['Taluka', 'Taluk', 'Mandal', 'Tehsil']  # Add more variations if needed
                matching_headers = set(taluka_headers) & set(headers)

                if matching_headers:
                    rows = table.find_all('tr')
                    for row in rows[1:]:
                        cells = row.find_all('td')
                        #taluka_name = cells[0].text.strip()
                        taluka_name = cells[0].text.strip().replace(" ", "-")  # Replace spaces with hyphens
                        taluka_names.append(taluka_name)
                    break  # Break the loop if taluka names are found

            return taluka_names

        taluka_names = extract_taluka_names(html)

        for taluka_name in taluka_names:
            url = f"https://www.censusindia2011.com/{STATE}/{formatted_district_name}/{taluka_name}-population.html"
            def get_html_inside_div(url, div_class):
                # Fetch the webpage content
                response = requests.get(url)

                # Check if the request was successful
                if response.status_code == 200:
                    # Parse the HTML content
                    soup = BeautifulSoup(response.content, 'html.parser')

                    # Find the div with the specified class
                    div = soup.find('div', class_=div_class)

                    # Check if the div is found
                    if div:
                        # Return the HTML content inside the div
                        return str(div)
                    else:
                        return "Div with class '{}' not found on the page.".format(div_class)
                else:
                    return "Failed to retrieve webpage. Status code: {}".format(response.status_code)
            html_inside_div = get_html_inside_div(url, 'mt20')
            html = str(html_inside_div)

            def html_to_table(html):
                soup = BeautifulSoup(html, 'html.parser')
                all_tables = soup.find_all('table')
                result = []
                for i, table in enumerate(all_tables):
                    headers = [header.text.strip() for header in table.find_all('th')]
                    if all(col_name in headers for col_name in ['Village', 'Population', 'Literacy', 'Sex-ratio']) or all(col_name in headers for col_name in ['Town', 'Population', 'Literacy', 'Sex-ratio']):
                        rows = []
                        for row in table.find_all('tr'):
                            cells = [cell.text.strip() for cell in row.find_all('td')]
                            if cells:
                                rows.append(cells)
                        result.append((headers, rows))
                return result

            tables = html_to_table(html)

            combined_sheet = wb.create_sheet(title=f'{taluka_name}')

            for s, (headers, rows) in enumerate(tables):
                df = pd.DataFrame(rows, columns=headers)

                # Append the data to the combined sheet
                for r_idx, row in enumerate(df.values, 1):
                    for c_idx, value in enumerate(row, 1):
                        combined_sheet.cell(row=r_idx, column=c_idx, value=value)

            # Remove the default sheet created by openpyxl (Sheet)
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # Save the workbook
        wb.save(os.path.join(output_directory, f'{district}.xlsx'))





st.title("GENERICART")
#st.write("If name of STATE contains more than 1 word then join with - ,like uttar-pradesh")
state_names = [
    'Uttar-Pradesh', 'Maharashtra', 'Bihar', 'West-Bengal', 'Andhra-Pradesh',
    'Madhya-Pradesh', 'Tamil-Nadu', 'Rajasthan', 'Karnataka', 'Gujarat',
    'Odisha', 'Kerala', 'Jharkhand', 'Assam', 'Punjab',
    'Haryana', 'NCT-Of-Delhi', 'Jammu-&-Kashmir', 'Uttarakhand',
    'Himachal-Pradesh', 'Tripura', 'Meghalaya', 'Manipur', 'Nagaland',
    'Goa', 'Arunachal-Pradesh', 'Puducherry', 'Mizoram', 'Chandigarh',
    'Sikkim', 'Andaman-&-Nicobar-Islands', 'Dadra-&-Nagar-Haveli',
    'Daman-&-Diu', 'Lakshadweep']

state=st.selectbox("Select a state:", state_names)

# Button to trigger the processing function
if st.button("Run Processing"):
    # Show a spinner while the processing is ongoing
    with st.spinner("Processing..."):
        # Call your time-consuming function here
        god(state)

    # Once the processing is done, remove the spinner
    st.success("Processing complete!")

st.subheader(f"Data Organization in Folder")

# Write information about the Excel file and data organization
st.write(
    f"In the D drive, folder has been identified, named after the specified state. "
    f"This file encompasses multiple sheets, each dedicated to individual districts within the state. "
    f"The organization of data within these sheets adheres to a structured format, where the sheet names correspond to the respective talukas."
)

# Display information about the columns in the DataFrame
st.write("Each sheet contains essential demographic information, meticulously arranged in columns for ease of analysis. The columns include:")
st.write("- **VILLAGE:** This field provides a comprehensive list of villages within the taluka, offering a granular view of the geographic distribution.")
st.write("- **POPULATION:** The population column quantifies the total number of residents in each village, facilitating an understanding of settlement sizes.")
st.write("- **LITERACY:** Literacy rates are documented, reflecting the proportion of individuals who possess basic reading and writing skills in the specified villages.")
st.write("- **SEX RATIO:** The sex ratio column provides insights into the gender distribution within the villages, aiding in the evaluation of gender demographics.")




